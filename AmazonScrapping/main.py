from bs4 import BeautifulSoup
import openpyxl


# Read the HTML file
with open('./amazon1.html', 'r',encoding= "utf8") as file:
    #opening file from local dir
    html_data = file.read()
    # print(html_data)

# Create BeautifulSoup object
soup = BeautifulSoup(html_data, 'html.parser')

# Find all divs with the specified class
divs = soup.find_all('div')
#div tag containing the product image,product rating, product name, product price etc .


# Create a workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write column headings for sheet
sheet.cell(row=1, column=1, value='Product_Name')
sheet.cell(row=1, column=2, value='Product_Price')
sheet.cell(row=1, column=3, value ='Bank_offers')


# Iterate over the divs and extract the required data
row = 2
#Since the column headings are written in the first row, the data extraction and writing begin from the second row onwards, allowing each product's information to be written in subsequent rows of the sheet.
for div in divs:
        for p in soup.find_all('span', attrs={'class': 'a-size-large product-title-word-break'}):
            product_name = p.text
            sheet.cell(row=row, column=1, value=product_name)
            print(p.text)
            for p in soup.find('span', attrs={'class': 'a-price aok-align-center reinventPricePriceToPayMargin priceToPay'}):
                product_price = p.text
            sheet.cell(row=row, column=2, value=product_price)
            print(p.text)


            for div in soup.find_all('div', class_="InstantBankDiscount-sideSheet"):
                for p in div.find_all('p', attrs={'class': 'a-spacing-mini a-size-base-plus'}):
                    Bank_offers = p.text
                    print(Bank_offers)
                    # Write the data to the Excel sheet


                    sheet.cell(row=row, column=3, value=Bank_offers)
                    # sheet.cell(row=row, column=4, value= product_Bank_offers)

                    row += 1
        break

# Save the workbook
workbook.save('amazon2.xlsx')
