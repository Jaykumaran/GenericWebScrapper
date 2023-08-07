from flask import Flask, render_template, request
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
from bs4 import BeautifulSoup

app = Flask(__name__)


def nykaa_extract_data(driver, product_name):
    # Find the search input field using its CSS class name
    search_input = driver.find_element(By.CSS_SELECTOR, '.css-1upamjb')

    # Enter the product name in the search input field
    search_input.send_keys(product_name)
    search_input.send_keys(Keys.ENTER)

    # Wait for the search results page to load
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.css-cjd9an')))

    # Find all product names and offer sections
    product_names = []
    offer_sections = []
    alternate_offers_list = []

    while True:
        # Find all product names on the current page
        product_name_div = driver.find_elements(By.CSS_SELECTOR, '.css-xrzmfa')
        product_names.extend([name.text for name in product_name_div])

        # Find all offer sections on the current page
        offers_span = driver.find_elements(By.CSS_SELECTOR, '.css-cjd9an')
        alternate_offers = driver.find_elements(By.CSS_SELECTOR, 'p.css-1kzcg63')

        for p in alternate_offers:
            offer_sections.append(p.text)

        for span in offers_span:
            try:
                if span.text:
                    offer_sections.append(span.text)
                else:
                    offer_sections.append('')
            except Exception as e:
                for offer in alternate_offers:
                    offer_sections.append(offer.text)

        # Check if "View More Products" button is present and click it to load more products
        view_more_button = driver.find_elements(By.XPATH, '//button[contains(@class, "load-more-button")]')
        if len(view_more_button) > 0:
            view_more_button[0].click()
            time.sleep(2)
        else:
            break

    # Create a workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write column headings for sheet
    sheet.cell(row=1, column=1, value='Product Name')
    sheet.cell(row=1, column=2, value='Offer Section')

    # Iterate over the product names and offer sections and write them to the Excel sheet
    row = 2
    for product_name, offer_section in zip(product_names, offer_sections):
        sheet.cell(row=row, column=1, value=product_name)
        sheet.cell(row=row, column=2, value=offer_section)
        row += 1

    # Save the workbook
    workbook.save('nyka.xlsx')

    # Print the product names and offers in the terminal
    for product_name, offer_section in zip(product_names, offer_sections):
        print("Product Name:", product_name)
        print("Offer Section:", offer_section)
        print()



def flipkart_extract_data(driver, product_name):
    """
    This function handles the login popup and extracts data from the Flipkart website.
    """
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, '_2KpZ6l._2doB4z')))
        login_close_button = driver.find_element(By.CLASS_NAME, '_2KpZ6l._2doB4z')
        login_close_button.click()
    except Exception as e:
        raise Exception("(login_popup_handle) - Failed to handle popup.\n" + str(e))

    # Find the search input field using its CSS class name
    search_input = driver.find_element(By.CLASS_NAME, '_3704LK')

    # Enter the product name in the search input field
    search_input.send_keys(product_name)
    search_input.send_keys(Keys.ENTER)

    # Wait for the search results page to load
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CLASS_NAME, '_4rR01T')))

    # Find all product names and offer sections
    product_names = []
    offer_sections = []

    # # Find all product names on the current page
    # Find all product names on the current page
    product_name_div = driver.find_elements(By.CLASS_NAME, '_4rR01T')

    # Loop through each product on the page
    for i in range(len(product_name_div)):
        # Click on the product to open it in a new tab
        product = product_name_div[i]
        time.sleep(2)
        # Open the product in a new tab
        actions = ActionChains(driver)
        actions.key_down(Keys.CONTROL).click(product).key_up(Keys.CONTROL).perform()

        # Switch to the newly opened tab
        driver.switch_to.window(driver.window_handles[1])
        driver.implicitly_wait(10)  # Add a small delay to let the new tab load
        view_offers_btn = driver.find_elements(By.CLASS_NAME, 'IMZJg1')
        if len(view_offers_btn) > 0:
            view_offers_btn[0].click()

        # Find the offer sections for the current product
        offers_span = driver.find_elements(By.CLASS_NAME, 'XUp0WS')

        for span in offers_span:
            offer_sections.append(span.text)

        # Close the current tab and switch back to the original tab
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        driver.switch_to.default_content()

        # Retrieve the product name and add it to the list
        product_name = product.text
        product_names.append(product_name)



    # Check if there is a "Next" button and click it to load more products
    next_button = driver.find_elements(By.CLASS_NAME, '_1LKTO3')
    if len(next_button) > 0:
        next_button[0].click()
        time.sleep(2)  # Add a small delay to let the next page load
    else:
        pass




    # Create a workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write column headings for sheet
    sheet.cell(row=1, column=1, value='Product Name')
    sheet.cell(row=1, column=2, value='Offer Section')

    # Iterate over the product names and offer sections and write them to the Excel sheet
    row = 2
    for product_name, offer_section in zip(product_names, offer_sections):
        sheet.cell(row=row, column=1, value=product_name)
        sheet.cell(row=row, column=2, value=offer_section)
        row += 1

    # Save the workbook
    workbook.save('flipkart1.xlsx')

    # Print the product names and offer sections
    for product_name, offer_section in zip(product_names, offer_sections):
        print("Product Name:", product_name)
        print("Offer Section:", offer_section)
        print()

def amazon_extract_data():



    # Read the HTML file
    with open('C:/Users/jaikr/PycharmProjects/GenericWebScrapper/AmazonScrapping/amazon1.html', 'r', encoding="utf8") as file:
        # opening file from local dir
        html_data = file.read()
        # print(html_data)

    # Create BeautifulSoup object
    soup = BeautifulSoup(html_data, 'html.parser')

    # Find all divs with the specified class
    divs = soup.find_all('div')
    # div tag containing the product image,product rating, product name, product price etc .

    # Create a workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write column headings for sheet
    sheet.cell(row=1, column=1, value='Product_Name')
    sheet.cell(row=1, column=2, value='Product_Price')
    sheet.cell(row=1, column=3, value='Bank_offers')

    # Iterate over the divs and extract the required data
    row = 2
    # Since the column headings are written in the first row, the data extraction and writing begin from the second row onwards, allowing each product's information to be written in subsequent rows of the sheet.
    for div in divs:
        for p in soup.find_all('span', attrs={'class': 'a-size-large product-title-word-break'}):
            product_name = p.text
            sheet.cell(row=row, column=1, value=product_name)
            print(p.text)
            for p in soup.find('span',
                               attrs={'class': 'a-price aok-align-center reinventPricePriceToPayMargin priceToPay'}):
                product_price = p.text
            sheet.cell(row=row, column=2, value=product_price)
            print(p.text)

            for div in soup.find_all('div', class_="InstantBankDiscount-sideSheet"):
                for p in div.find_all('p', attrs={'class': 'a-spacing-mini a-size-base-plus'}):
                    Bank_offers = p.text
                    print(Bank_offers)
                    # Write the data to the Excel sheet

                    sheet.cell(row=row, column=3, value=Bank_offers)
                    # sheet.cell(row=row, column=4, value= product_Bank_offers)

                    row += 1
        break

    # Save the workbook
    workbook.save('amazon1_data.xlsx')


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/scrape', methods=['POST'])
def scrape():
    url = request.form['url']
    products_name = request.form['product_name']
    # Check the URL and set the product_name accordingly

    if url == 'nykaa':
        url= 'https://www.nykaa.com/'


        # Create a Selenium driver and navigate to the provided URL
        driver = webdriver.Chrome()
        driver.get(url)

        # Target the search column and enter a product name like "nykaa naturals"
        product_name = products_name

        nykaa_extract_data(driver, product_name)

        # Close the Selenium driver
        driver.quit()

        return 'Scraping completed.'
    elif url == 'flipkart':

        url = 'https://www.flipkart.com/'


        # Create a Selenium driver and navigate to the provided URL
        driver = webdriver.Chrome()
        driver.get(url)

        # Target the search column and enter a product name like "nykaa naturals"
        product_name = products_name

        flipkart_extract_data(driver, product_name)

        # Close the Selenium driver
        driver.quit()

        return 'Scraping completed.'
    elif url == 'amazon':
        amazon_extract_data()




if __name__ == '__main__':
    app.run()


